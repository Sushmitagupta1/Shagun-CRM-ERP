import uuid
import math
from datetime import date, datetime, time
from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, or_, update
from app.database import get_db
from app.models.inquiry import Inquiry, InquiryStatus, PaymentStatus, FollowUp, Meeting
from app.models.inventory_movement import InventoryMovement
from app.models.notification import Notification
from app.models.menu_slot import MenuSlot
from app.models.user import User, Role
from app.schemas.inquiry import InquiryCreate, InquiryUpdate, InquiryResponse, FollowUpCreate, FollowUpUpdate, FollowUpResponse, MeetingCreate, MeetingStatusUpdate, MeetingResponse, MenuSlotResponse, CalendarResponse
from app.schemas.inventory import InventoryMovementCreate, InventoryMovementResponse
from app.schemas.common import PaginatedResponse
from app.middleware.auth import get_current_user

from app.services.inquiry_service import get_inquiry_or_404
import os
from fastapi import UploadFile, File
from app.config import settings

router = APIRouter(prefix="/api/inquiries", tags=["inquiries"])

FOLLOWUP_NOTIFY_ROLES = ("admin", "sales_head", "presentation_exec")
FOLLOWUP_WRITE_ROLES = ("admin", "sales_head", "presentation_exec")


def apply_filters(query, count_query, status, assigned_to, search, event_type, date_from, date_to, followup=None):
    if status:
        statuses = [s.strip() for s in status.split(",") if s.strip()]
        if len(statuses) > 1:
            query = query.where(Inquiry.status.in_(statuses))
            count_query = count_query.where(Inquiry.status.in_(statuses))
        else:
            query = query.where(Inquiry.status == status)
            count_query = count_query.where(Inquiry.status == status)
    if assigned_to:
        query = query.where(Inquiry.assigned_to == assigned_to)
        count_query = count_query.where(Inquiry.assigned_to == assigned_to)
    if event_type:
        query = query.where(Inquiry.event_type.ilike(f"%{event_type}%"))
        count_query = count_query.where(Inquiry.event_type.ilike(f"%{event_type}%"))
    if search:
        search_filter = or_(Inquiry.client_name.ilike(f"%{search}%"), Inquiry.client_phone.ilike(f"%{search}%"))
        query = query.where(search_filter)
        count_query = count_query.where(search_filter)
    if date_from:
        query = query.where(Inquiry.inquiry_date >= date.fromisoformat(date_from))
        count_query = count_query.where(Inquiry.inquiry_date >= date.fromisoformat(date_from))
    if date_to:
        query = query.where(Inquiry.inquiry_date <= date.fromisoformat(date_to))
        count_query = count_query.where(Inquiry.inquiry_date <= date.fromisoformat(date_to))
    if followup:
        today = date.today()
        if followup == "today":
            subq = select(FollowUp.inquiry_id).where(
                FollowUp.follow_up_date == today, FollowUp.is_done.is_(False)
            )
        elif followup == "upcoming":
            subq = select(FollowUp.inquiry_id).where(
                FollowUp.follow_up_date >= today, FollowUp.is_done.is_(False)
            )
        elif followup == "overdue":
            subq = (
                select(FollowUp.inquiry_id)
                .join(Inquiry, FollowUp.inquiry_id == Inquiry.id)
                .where(FollowUp.follow_up_date < today, FollowUp.is_done.is_(False),
                       Inquiry.status.in_([InquiryStatus.NEW_INQUIRY, InquiryStatus.FOLLOWUP]))
            )
        else:
            subq = None
        if subq is not None:
            query = query.where(Inquiry.id.in_(subq))
            count_query = count_query.where(Inquiry.id.in_(subq))
    return query, count_query


@router.get("", response_model=PaginatedResponse[InquiryResponse])
async def list_inquiries(
    page: int = Query(1, ge=1), per_page: int = Query(20, ge=1, le=100),
    status: str | None = None, assigned_to: uuid.UUID | None = None,
    search: str | None = None, event_type: str | None = None,
    date_from: str | None = None, date_to: str | None = None,
    event_date_from: str | None = None, event_date_to: str | None = None,
    followup: str | None = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    query = select(Inquiry)
    count_query = select(func.count(Inquiry.id))
    query, count_query = apply_filters(query, count_query, status, assigned_to, search, event_type, date_from, date_to, followup)
    if event_date_from:
        query = query.where(Inquiry.event_date >= date.fromisoformat(event_date_from))
        count_query = count_query.where(Inquiry.event_date >= date.fromisoformat(event_date_from))
    if event_date_to:
        query = query.where(Inquiry.event_date <= date.fromisoformat(event_date_to))
        count_query = count_query.where(Inquiry.event_date <= date.fromisoformat(event_date_to))
    total_result = await db.execute(count_query)
    total = total_result.scalar()
    query = query.order_by(Inquiry.created_at.desc()).offset((page - 1) * per_page).limit(per_page)
    result = await db.execute(query)
    inquiries = result.scalars().all()

    # Populate the next pending follow-up date for each inquiry in this page.
    next_follow_up_map: dict = {}
    if inquiries:
        ids = [i.id for i in inquiries]
        fu_result = await db.execute(
            select(FollowUp.inquiry_id, func.min(FollowUp.follow_up_date))
            .where(FollowUp.inquiry_id.in_(ids), FollowUp.is_done.is_(False))
            .group_by(FollowUp.inquiry_id)
        )
        next_follow_up_map = dict(fu_result.all())
    items = []
    for i in inquiries:
        resp = InquiryResponse.model_validate(i)
        resp.next_follow_up = next_follow_up_map.get(i.id)
        items.append(resp)

    return PaginatedResponse(
        items=items,
        total=total, page=page, per_page=per_page,
        total_pages=math.ceil(total / per_page) if total > 0 else 0,
    )


@router.get("/calendar", response_model=CalendarResponse)
async def calendar_events(
    from_date: str = Query(...), to_date: str = Query(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    start = date.fromisoformat(from_date)
    end = date.fromisoformat(to_date)
    events_result = await db.execute(
        select(Inquiry).where(Inquiry.event_date.isnot(None), Inquiry.event_date >= start, Inquiry.event_date <= end)
    )
    events = [InquiryResponse.model_validate(i) for i in events_result.scalars().all()]

    fu_result = await db.execute(
        select(FollowUp, Inquiry.client_name, Inquiry.event_type)
        .join(Inquiry, FollowUp.inquiry_id == Inquiry.id)
        .where(FollowUp.follow_up_date >= start, FollowUp.follow_up_date <= end)
        .order_by(FollowUp.follow_up_date.asc())
    )
    followups = [
        {
            "id": str(fu.id),
            "inquiry_id": str(fu.inquiry_id),
            "client_name": name,
            "event_type": event_type,
            "follow_up_date": fu.follow_up_date.isoformat(),
            "remarks": fu.remarks,
            "is_done": fu.is_done,
        }
        for fu, name, event_type in fu_result.all()
    ]

    m_result = await db.execute(
        select(Meeting, Inquiry.client_name, Inquiry.event_type)
        .join(Inquiry, Meeting.inquiry_id == Inquiry.id)
        .where(Meeting.meeting_at >= datetime.combine(start, time.min), Meeting.meeting_at <= datetime.combine(end, time.max))
        .order_by(Meeting.meeting_at.asc())
    )
    meetings = [
        {
            "id": str(m.id),
            "inquiry_id": str(m.inquiry_id),
            "client_name": name,
            "event_type": event_type,
            "meeting_at": m.meeting_at.isoformat(),
            "remarks": m.remarks,
            "status": m.status,
        }
        for m, name, event_type in m_result.all()
    ]

    return CalendarResponse(events=events, followups=followups, meetings=meetings)


@router.get("/export/excel")
async def export_inquiries_excel(
    status: str | None = None, assigned_to: uuid.UUID | None = None,
    search: str | None = None, event_type: str | None = None,
    date_from: str | None = None, date_to: str | None = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from openpyxl import Workbook
    from io import BytesIO

    query = select(Inquiry)
    count_query = select(func.count(Inquiry.id))
    query, _ = apply_filters(query, count_query, status, assigned_to, search, event_type, date_from, date_to)
    query = query.order_by(Inquiry.created_at.desc())
    result = await db.execute(query)
    inquiries = result.scalars().all()

    next_follow_up_map: dict = {}
    if inquiries:
        ids = [i.id for i in inquiries]
        fu_result = await db.execute(
            select(FollowUp.inquiry_id, func.min(FollowUp.follow_up_date))
            .where(FollowUp.inquiry_id.in_(ids), FollowUp.is_done.is_(False))
            .group_by(FollowUp.inquiry_id)
        )
        next_follow_up_map = dict(fu_result.all())

    wb = Workbook()
    ws = wb.active
    ws.title = "Inquiries"
    ws.append(["Event Date", "Client Name", "Phone", "Event Type", "Pax", "Venue",
               "Inquiry Date", "Follow-up Date", "Status"])

    for i in inquiries:
        ws.append([
            i.event_date.isoformat() if i.event_date else "",
            i.client_name, i.client_phone, i.event_type, i.pax, i.venue or "",
            i.inquiry_date.isoformat() if i.inquiry_date else "",
            next_follow_up_map.get(i.id).isoformat() if next_follow_up_map.get(i.id) else "",
            i.status.value if hasattr(i.status, 'value') else i.status,
        ])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=inquiries.xlsx"},
    )


@router.get("/{inquiry_id}/export/excel")
async def export_single_inquiry_excel(
    inquiry_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from openpyxl import Workbook
    from io import BytesIO
    from fastapi.responses import StreamingResponse

    inquiry = await get_inquiry_or_404(db, inquiry_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Inquiry"
    ws.append(["Field", "Value"])
    ws.append(["Client Name", inquiry.client_name])
    ws.append(["Phone", inquiry.client_phone])
    ws.append(["Event Type", inquiry.event_type])
    ws.append(["Pax", inquiry.pax])
    ws.append(["Per Plate Rate", float(inquiry.per_plate_rate or 0)])
    ws.append(["Add On", float(inquiry.add_on or 0)])
    ws.append(["Total Amount", (float(inquiry.per_plate_rate or 0) * (inquiry.pax or 0)) + float(inquiry.add_on or 0)])
    ws.append(["Inquiry Date", inquiry.inquiry_date.isoformat() if inquiry.inquiry_date else ""])
    ws.append(["Event Date", inquiry.event_date.isoformat() if inquiry.event_date else ""])
    ws.append(["Status", inquiry.status.value if hasattr(inquiry.status, 'value') else inquiry.status])
    ws.append(["Payment Status", inquiry.payment_status.value if hasattr(inquiry.payment_status, 'value') else inquiry.payment_status])
    ws.append(["Advance Amount", float(inquiry.advance_amount)])
    ws.append(["Remarks", inquiry.remarks or ""])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=inquiry_{inquiry.client_name}.xlsx"},
    )


@router.get("/{inquiry_id}", response_model=InquiryResponse)
async def get_inquiry(inquiry_id: uuid.UUID, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    inquiry = await get_inquiry_or_404(db, inquiry_id)
    return InquiryResponse.model_validate(inquiry)


@router.get("/{inquiry_id}/menu/download")
async def download_menu(inquiry_id: uuid.UUID, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    inquiry = await get_inquiry_or_404(db, inquiry_id)
    if not inquiry.menu_content:
        raise HTTPException(status_code=404, detail="No menu content available")
    from fastapi.responses import Response
    filename = f"Menu_{inquiry.client_name.replace(' ', '_')}.txt"
    return Response(
        content=inquiry.menu_content,
        media_type="text/plain",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


ALLOWED_ROLES = {
    "menu": {"admin", "menu_planner"},
    "presentation": {"admin", "presentation_exec"},
    "ingredient": {"admin", "kitchen"},
    "inventory": {"admin", "operations_manager", "warehouse"},
    "returned": {"admin", "operations_manager", "warehouse"},
    "transferred": {"admin", "operations_manager", "warehouse"},
    "wastage": {"admin", "operations_manager", "warehouse"},
    "call_recording": {"admin", "sales_head", "presentation_exec"},
}

FILE_TYPES = tuple(ALLOWED_ROLES.keys())

INVENTORY_FILE_COLUMNS = {
    "received": ("inventory_file_name", "inventory_file_path"),
    "returned": ("returned_file_name", "returned_file_path"),
    "transferred": ("transferred_file_name", "transferred_file_path"),
    "wastage": ("wastage_file_name", "wastage_file_path"),
}


@router.post("/{inquiry_id}/upload")
async def upload_inquiry_file(
    inquiry_id: uuid.UUID,
    file_type: str = Query(...),
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if file_type not in FILE_TYPES:
        raise HTTPException(status_code=400, detail=f"file_type must be one of: {', '.join(FILE_TYPES)}")
    if current_user.role.name not in ALLOWED_ROLES[file_type]:
        raise HTTPException(status_code=403, detail="Not authorized")
    inquiry = await get_inquiry_or_404(db, inquiry_id)

    ext = os.path.splitext(file.filename or "")[1].lower()
    if ext not in settings.ALLOWED_EXTENSIONS:
        raise HTTPException(status_code=400, detail=f"File type '{ext}' not allowed")
    content = await file.read()
    max_size = settings.MAX_CALL_RECORDING_SIZE if file_type == "call_recording" else settings.MAX_UPLOAD_SIZE
    if len(content) > max_size:
        raise HTTPException(status_code=400, detail=f"File too large (max {max_size // (1024 * 1024)}MB)")

    upload_dir = os.path.join(settings.UPLOAD_DIR, str(inquiry_id), file_type)
    os.makedirs(upload_dir, exist_ok=True)
    file_path = os.path.join(upload_dir, file.filename or "unnamed")
    with open(file_path, "wb") as f:
        f.write(content)

    setattr(inquiry, f"{file_type}_file_name", file.filename)
    setattr(inquiry, f"{file_type}_file_path", file_path)
    await db.commit()
    await db.refresh(inquiry)

    if file_type in ("menu", "presentation"):
        notify_result = await db.execute(
            select(User)
            .join(User.role)
            .where(Role.name == "sales_head", User.is_active.is_(True))
        )
        notify_users = notify_result.scalars().all()
        label = "Menu" if file_type == "menu" else "Presentation"
        message = f"{label} uploaded for {inquiry.client_name} ({inquiry.event_type})"
        for u in notify_users:
            db.add(Notification(
                user_id=u.id,
                title=f"{label} uploaded",
                message=message[:500],
                type="file_upload",
                entity_type="inquiry",
                entity_id=inquiry.id,
            ))
        await db.commit()

    return {"file_name": file.filename, "file_path": file_path}


INVENTORY_HEADER_WORDS = {"item", "item name", "item_name", "itemname", "product", "material", "ingredient", "name", "description"}


def parse_movement_file(file_path: str, ext: str):
    rows: list[tuple[str, float, str | None]] = []
    if ext == ".csv":
        import csv
        with open(file_path, "r", encoding="utf-8", errors="replace") as f:
            reader = csv.reader(f)
            for raw in reader:
                row = [c.strip() for c in raw]
                if not any(row):
                    continue
                item = row[0] if len(row) > 0 else ""
                if not item or item.lower() in INVENTORY_HEADER_WORDS:
                    continue
                try:
                    qty = float((row[1] if len(row) > 1 else "").replace(",", "")) if len(row) > 1 and row[1] else 0.0
                except (ValueError, TypeError):
                    qty = 0.0
                if qty <= 0:
                    continue
                unit = row[2] if len(row) > 2 and row[2] else None
                rows.append((item, qty, unit))
    else:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        for raw in ws.iter_rows(values_only=True):
            row = ["" if c is None else str(c).strip() for c in raw]
            if not any(row):
                continue
            item = row[0] if len(row) > 0 else ""
            if not item or item.lower() in INVENTORY_HEADER_WORDS:
                continue
            try:
                qty = float((row[1] if len(row) > 1 else "").replace(",", "")) if len(row) > 1 and row[1] else 0.0
            except (ValueError, TypeError):
                qty = 0.0
            if qty <= 0:
                continue
            unit = row[2] if len(row) > 2 and row[2] else None
            rows.append((item, qty, unit))
    return rows


@router.post("/{inquiry_id}/inventory-upload")
async def upload_inventory_movement_file(
    inquiry_id: uuid.UUID,
    movement_type: str = Query(...),
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if movement_type not in INVENTORY_FILE_COLUMNS:
        raise HTTPException(status_code=400, detail="movement_type must be one of: received, returned, transferred, wastage")
    if current_user.role.name not in ("admin", "operations_manager", "warehouse"):
        raise HTTPException(status_code=403, detail="Not authorized")
    inquiry = await get_inquiry_or_404(db, inquiry_id)

    ext = os.path.splitext(file.filename or "")[1].lower()
    if ext not in (".xlsx", ".csv"):
        raise HTTPException(status_code=400, detail="Only .xlsx or .csv files are supported")
    content = await file.read()
    if len(content) > settings.MAX_UPLOAD_SIZE:
        raise HTTPException(status_code=400, detail="File too large (max 20MB)")

    upload_dir = os.path.join(settings.UPLOAD_DIR, str(inquiry_id), "inventory", movement_type)
    os.makedirs(upload_dir, exist_ok=True)
    file_path = os.path.join(upload_dir, file.filename or "unnamed")
    with open(file_path, "wb") as f:
        f.write(content)

    rows = parse_movement_file(file_path, ext)

    old = await db.execute(select(InventoryMovement).where(InventoryMovement.inquiry_id == inquiry_id, InventoryMovement.movement_type == movement_type))
    for m in old.scalars().all():
        await db.delete(m)
    for item, qty, unit in rows:
        db.add(InventoryMovement(
            inquiry_id=inquiry_id,
            movement_type=movement_type,
            item_name=item,
            quantity=qty,
            unit=unit,
            created_by=current_user.id,
        ))

    name_col, path_col = INVENTORY_FILE_COLUMNS[movement_type]
    setattr(inquiry, name_col, file.filename)
    setattr(inquiry, path_col, file_path)
    await db.commit()

    return {"file_name": file.filename, "entries_created": len(rows)}


MAX_PREVIEW_ROWS = 200
MAX_PREVIEW_COLS = 12


def read_file_preview(file_path: str, ext: str):
    rows: list[list] = []
    if ext == ".csv":
        import csv
        with open(file_path, "r", encoding="utf-8", errors="replace") as f:
            reader = csv.reader(f)
            for raw in reader:
                if len(rows) >= MAX_PREVIEW_ROWS:
                    break
                row = ["" if c is None else str(c).strip() for c in raw[:MAX_PREVIEW_COLS]]
                if any(row):
                    rows.append(row)
    else:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, data_only=True, read_only=True)
        try:
            ws = wb.active
            for raw in ws.iter_rows(values_only=True):
                if len(rows) >= MAX_PREVIEW_ROWS:
                    break
                row = ["" if c is None else (c if isinstance(c, (int, float)) else str(c).strip()) for c in raw[:MAX_PREVIEW_COLS]]
                if any(row):
                    rows.append(row)
        finally:
            wb.close()
    return rows


@router.get("/{inquiry_id}/file/{file_type}/preview")
async def preview_inquiry_file(
    inquiry_id: uuid.UUID,
    file_type: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if file_type not in FILE_TYPES:
        raise HTTPException(status_code=400, detail=f"file_type must be one of: {', '.join(FILE_TYPES)}")
    inquiry = await get_inquiry_or_404(db, inquiry_id)
    file_path = getattr(inquiry, f"{file_type}_file_path", None)
    file_name = getattr(inquiry, f"{file_type}_file_name", None)
    if not file_path or not os.path.isfile(file_path):
        raise HTTPException(status_code=404, detail="No file uploaded")
    ext = os.path.splitext(file_name or "")[1].lower()
    if ext not in (".xlsx", ".csv"):
        raise HTTPException(status_code=400, detail="Preview is available for .xlsx and .csv files only")
    rows = read_file_preview(file_path, ext)
    return {"file_name": file_name, "rows": rows}


@router.get("/{inquiry_id}/file/{file_type}")
async def download_inquiry_file(
    inquiry_id: uuid.UUID,
    file_type: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if file_type not in FILE_TYPES:
        raise HTTPException(status_code=400, detail=f"file_type must be one of: {', '.join(FILE_TYPES)}")
    inquiry = await get_inquiry_or_404(db, inquiry_id)
    file_path = getattr(inquiry, f"{file_type}_file_path", None)
    file_name = getattr(inquiry, f"{file_type}_file_name", None)
    if not file_path or not os.path.isfile(file_path):
        raise HTTPException(status_code=404, detail="No file uploaded")
    from fastapi.responses import FileResponse
    return FileResponse(
        path=file_path,
        filename=file_name,
        headers={"Cache-Control": "no-store, no-cache, must-revalidate"},
    )


@router.post("", response_model=InquiryResponse, status_code=201)
async def create_inquiry(data: InquiryCreate, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    inquiry = Inquiry(
        id=uuid.uuid4(), client_name=data.client_name, client_phone=data.client_phone,
        event_type=data.event_type, session=data.session, source=data.source,
        event_date=data.event_date, pax=data.pax,
        inquiry_date=data.inquiry_date or date.today(),
        per_plate_rate=data.per_plate_rate, add_on=data.add_on,
        assigned_to=data.assigned_to, remarks=data.remarks, venue=data.venue,
        created_by=current_user.id,
        status=InquiryStatus.FOLLOWUP if data.follow_up_date else InquiryStatus.NEW_INQUIRY,
        payment_status=PaymentStatus.UNPAID,
    )
    db.add(inquiry)
    await db.flush()

    if data.follow_up_date:
        follow_up = FollowUp(
            id=uuid.uuid4(), inquiry_id=inquiry.id,
            follow_up_date=data.follow_up_date,
            created_by=current_user.id,
        )
        db.add(follow_up)

    await db.commit()
    await db.refresh(inquiry)
    return InquiryResponse.model_validate(inquiry)


@router.put("/{inquiry_id}", response_model=InquiryResponse)
async def update_inquiry(inquiry_id: uuid.UUID, data: InquiryUpdate, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    inquiry = await get_inquiry_or_404(db, inquiry_id)
    for field, value in data.model_dump(exclude_unset=True).items():
        setattr(inquiry, field, value)
    await db.flush()
    await db.commit()
    await db.refresh(inquiry)
    return InquiryResponse.model_validate(inquiry)


@router.get("/{inquiry_id}/follow-ups", response_model=list[FollowUpResponse])
async def list_follow_ups(
    inquiry_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    result = await db.execute(
        select(FollowUp)
        .where(FollowUp.inquiry_id == inquiry_id)
        .order_by(FollowUp.follow_up_date.asc())
    )
    return [FollowUpResponse.model_validate(fu) for fu in result.scalars().all()]


@router.post("/{inquiry_id}/follow-ups", response_model=FollowUpResponse, status_code=201)
async def add_follow_up(
    inquiry_id: uuid.UUID, data: FollowUpCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    inquiry = await get_inquiry_or_404(db, inquiry_id)
    if current_user.role.name not in FOLLOWUP_WRITE_ROLES:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not authorized")
    follow_up = FollowUp(
        id=uuid.uuid4(), inquiry_id=inquiry.id,
        follow_up_date=data.follow_up_date,
        remarks=data.remarks,
        created_by=current_user.id,
    )
    if inquiry.status == InquiryStatus.NEW_INQUIRY:
        inquiry.status = InquiryStatus.FOLLOWUP
    db.add(follow_up)
    notify_result = await db.execute(
        select(User)
        .join(User.role)
        .where(Role.name.in_(FOLLOWUP_NOTIFY_ROLES), User.is_active.is_(True))
    )
    notify_users = notify_result.scalars().all()
    message = f"Follow-up scheduled for {inquiry.client_name} ({inquiry.event_type}) on {data.follow_up_date.strftime('%d %b %Y')}"
    if data.remarks:
        message += f" — {data.remarks}"
    message = message[:500]
    for u in notify_users:
        db.add(Notification(
            user_id=u.id,
            title="New follow-up",
            message=message,
            type="followup",
            entity_type="inquiry",
            entity_id=inquiry.id,
        ))
    await db.commit()
    await db.refresh(follow_up)
    return FollowUpResponse.model_validate(follow_up)


@router.patch("/{inquiry_id}/follow-ups/{follow_up_id}", response_model=FollowUpResponse)
async def update_follow_up(
    inquiry_id: uuid.UUID, follow_up_id: uuid.UUID, data: FollowUpUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    result = await db.execute(
        select(FollowUp).where(FollowUp.id == follow_up_id, FollowUp.inquiry_id == inquiry_id)
    )
    follow_up = result.scalar_one_or_none()
    if not follow_up:
        raise HTTPException(status_code=404, detail="Follow-up not found")
    if current_user.role.name not in FOLLOWUP_WRITE_ROLES:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not authorized")
    if data.is_done and not (data.remarks or "").strip():
        raise HTTPException(status_code=400, detail="Remark is required when marking a follow-up as done")
    follow_up.is_done = data.is_done
    if data.remarks is not None:
        follow_up.remarks = data.remarks
    await db.commit()
    await db.refresh(follow_up)
    return FollowUpResponse.model_validate(follow_up)


@router.get("/{inquiry_id}/meetings", response_model=list[MeetingResponse])
async def list_meetings(
    inquiry_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    await get_inquiry_or_404(db, inquiry_id)
    result = await db.execute(
        select(Meeting)
        .where(Meeting.inquiry_id == inquiry_id)
        .order_by(Meeting.meeting_at.asc())
    )
    return [MeetingResponse.model_validate(m) for m in result.scalars().all()]


@router.post("/{inquiry_id}/meetings", response_model=MeetingResponse, status_code=201)
async def add_meeting(
    inquiry_id: uuid.UUID, data: MeetingCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    inquiry = await get_inquiry_or_404(db, inquiry_id)
    meeting = Meeting(
        id=uuid.uuid4(), inquiry_id=inquiry.id,
        meeting_at=data.meeting_at,
        remarks=data.remarks,
        created_by=current_user.id,
    )
    db.add(meeting)
    notify_result = await db.execute(
        select(User)
        .join(User.role)
        .where(Role.name == "sales_head", User.is_active.is_(True))
    )
    notify_users = notify_result.scalars().all()
    message = f"Meeting scheduled for {inquiry.client_name} ({inquiry.event_type}) on {data.meeting_at.strftime('%d %b %Y, %I:%M %p')}"
    if data.remarks:
        message += f" — {data.remarks}"
    message = message[:500]
    for u in notify_users:
        db.add(Notification(
            user_id=u.id,
            title="New meeting",
            message=message,
            type="meeting",
            entity_type="inquiry",
            entity_id=inquiry.id,
        ))
    await db.commit()
    await db.refresh(meeting)
    return MeetingResponse.model_validate(meeting)


@router.patch("/{inquiry_id}/meetings/{meeting_id}", response_model=MeetingResponse)
async def update_meeting_status(
    inquiry_id: uuid.UUID, meeting_id: uuid.UUID, data: MeetingStatusUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    result = await db.execute(
        select(Meeting).where(Meeting.id == meeting_id, Meeting.inquiry_id == inquiry_id)
    )
    meeting = result.scalar_one_or_none()
    if not meeting:
        raise HTTPException(status_code=404, detail="Meeting not found")
    if data.status == "completed" and not (data.remarks or "").strip():
        raise HTTPException(status_code=400, detail="Remark is required when marking a meeting as complete")
    meeting.status = data.status
    if data.remarks is not None:
        meeting.remarks = data.remarks.strip()
    if data.status == "completed":
        inquiry = await get_inquiry_or_404(db, inquiry_id)
        notify_result = await db.execute(
            select(User)
            .join(User.role)
            .where(Role.name == "sales_head", User.is_active.is_(True))
        )
        notify_users = notify_result.scalars().all()
        message = f"Meeting completed for {inquiry.client_name} ({inquiry.event_type})"
        if meeting.remarks:
            message += f" — {meeting.remarks}"
        message = message[:500]
        for u in notify_users:
            db.add(Notification(
                user_id=u.id,
                title="Meeting completed",
                message=message,
                type="meeting",
                entity_type="inquiry",
                entity_id=inquiry.id,
            ))
    await db.commit()
    await db.refresh(meeting)
    return MeetingResponse.model_validate(meeting)


INVENTORY_MOVEMENT_WRITE_ROLES = ("admin", "operations_manager", "warehouse")


@router.get("/{inquiry_id}/inventory-movements", response_model=list[InventoryMovementResponse])
async def list_inventory_movements(
    inquiry_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    await get_inquiry_or_404(db, inquiry_id)
    result = await db.execute(
        select(InventoryMovement)
        .where(InventoryMovement.inquiry_id == inquiry_id)
        .order_by(InventoryMovement.created_at.desc())
    )
    return [InventoryMovementResponse.model_validate(m) for m in result.scalars().all()]


@router.post("/{inquiry_id}/inventory-movements", response_model=InventoryMovementResponse, status_code=201)
async def add_inventory_movement(
    inquiry_id: uuid.UUID, data: InventoryMovementCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role.name not in INVENTORY_MOVEMENT_WRITE_ROLES:
        raise HTTPException(status_code=403, detail="Not authorized")
    await get_inquiry_or_404(db, inquiry_id)
    movement = InventoryMovement(
        inquiry_id=inquiry_id,
        movement_type=data.movement_type,
        item_name=data.item_name.strip(),
        quantity=data.quantity,
        unit=data.unit,
        notes=data.notes,
        created_by=current_user.id,
    )
    db.add(movement)
    await db.commit()
    await db.refresh(movement)
    return InventoryMovementResponse.model_validate(movement)


@router.delete("/{inquiry_id}/inventory-movements/{movement_id}")
async def delete_inventory_movement(
    inquiry_id: uuid.UUID, movement_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role.name not in INVENTORY_MOVEMENT_WRITE_ROLES:
        raise HTTPException(status_code=403, detail="Not authorized")
    result = await db.execute(
        select(InventoryMovement).where(
            InventoryMovement.id == movement_id,
            InventoryMovement.inquiry_id == inquiry_id,
        )
    )
    movement = result.scalar_one_or_none()
    if not movement:
        raise HTTPException(status_code=404, detail="Movement entry not found")
    await db.delete(movement)
    await db.commit()
    return {"message": "Movement entry deleted"}


@router.patch("/{inquiry_id}/status")
async def update_status(inquiry_id: uuid.UUID, new_status: str, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    if current_user.role.name not in ("admin", "sales_head", "presentation_exec"):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Only admin, sales head or presentation exec can update status")
    inquiry = await get_inquiry_or_404(db, inquiry_id)
    try:
        target_status = InquiryStatus(new_status)
    except ValueError:
        raise HTTPException(status_code=400, detail=f"Invalid status: {new_status}")
    inquiry.status = target_status
    await db.flush()
    return {"message": f"Status updated to {target_status.value}"}


@router.patch("/{inquiry_id}/presentation-not-required")
async def update_presentation_not_required(
    inquiry_id: uuid.UUID,
    not_required: bool = Query(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role.name not in ("admin", "presentation_exec"):
        raise HTTPException(status_code=403, detail="Not authorized")
    inquiry = await get_inquiry_or_404(db, inquiry_id)
    inquiry.presentation_not_required = not_required
    await db.flush()
    return {"message": "Presentation marked as not required" if not_required else "Presentation marked as required"}


@router.patch("/{inquiry_id}/payment")
async def update_payment(inquiry_id: uuid.UUID, payment_status: str, advance_amount: float | None = None, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    inquiry = await get_inquiry_or_404(db, inquiry_id)
    if inquiry.status not in (InquiryStatus.ADVANCE_RECEIVE, InquiryStatus.OPERATION_HANDOVER):
        raise HTTPException(status_code=400, detail="Payment can only be updated after advance receive")
    inquiry.payment_status = PaymentStatus(payment_status)
    if advance_amount is not None:
        from decimal import Decimal
        inquiry.advance_amount = Decimal(str(advance_amount))
    await db.flush()
    return {"message": f"Payment status updated to {payment_status}"}


@router.patch("/{inquiry_id}/approve-payment")
async def approve_payment(inquiry_id: uuid.UUID, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    if current_user.role.name != "admin":
        raise HTTPException(status_code=403, detail="Only admin can approve payments")
    inquiry = await get_inquiry_or_404(db, inquiry_id)
    if inquiry.status not in (
        InquiryStatus.CLIENT_CONFIRMATION,
        InquiryStatus.ADVANCE_RECEIVE,
        InquiryStatus.OPERATION_HANDOVER,
    ):
        raise HTTPException(status_code=400, detail="Payment can only be approved at confirmation stage")
    inquiry.payment_status = PaymentStatus.PAID
    if inquiry.advance_payment_date is None:
        inquiry.advance_payment_date = date.today()
    if inquiry.status == InquiryStatus.CLIENT_CONFIRMATION:
        inquiry.status = InquiryStatus.ADVANCE_RECEIVE
    await db.flush()
    notify_result = await db.execute(
        select(User)
        .join(User.role)
        .where(Role.name == "sales_head", User.is_active.is_(True))
    )
    message = f"Payment approved for {inquiry.client_name} ({inquiry.event_type}) — {inquiry.client_phone or ''}"
    for u in notify_result.scalars().all():
        db.add(Notification(
            user_id=u.id,
            title="Payment approved",
            message=message[:500],
            type="payment",
            entity_type="inquiry",
            entity_id=inquiry.id,
        ))
    await db.commit()
    return {"message": "Payment approved"}


MENU_SLOT_WRITE_ROLES = ("admin", "menu_planner")


@router.get("/{inquiry_id}/menu-slots", response_model=list[MenuSlotResponse])
async def list_menu_slots(
    inquiry_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    await get_inquiry_or_404(db, inquiry_id)
    result = await db.execute(
        select(MenuSlot)
        .where(MenuSlot.inquiry_id == inquiry_id)
        .order_by(MenuSlot.slot_number.asc())
    )
    return [MenuSlotResponse.model_validate(s) for s in result.scalars().all()]


@router.post("/{inquiry_id}/menu-slots", response_model=MenuSlotResponse, status_code=201)
async def create_menu_slot(
    inquiry_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role.name not in MENU_SLOT_WRITE_ROLES:
        raise HTTPException(status_code=403, detail="Not authorized")
    await get_inquiry_or_404(db, inquiry_id)
    max_result = await db.execute(
        select(func.max(MenuSlot.slot_number)).where(MenuSlot.inquiry_id == inquiry_id)
    )
    next_number = (max_result.scalar() or 0) + 1
    if next_number > 7:
        raise HTTPException(status_code=400, detail="Maximum 7 menu slots allowed")
    slot = MenuSlot(
        inquiry_id=inquiry_id,
        slot_number=next_number,
        created_by=current_user.id,
    )
    db.add(slot)
    await db.commit()
    await db.refresh(slot)
    return MenuSlotResponse.model_validate(slot)


@router.post("/{inquiry_id}/menu-slots/{slot_id}/upload", response_model=MenuSlotResponse)
async def upload_menu_slot_file(
    inquiry_id: uuid.UUID,
    slot_id: uuid.UUID,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role.name not in MENU_SLOT_WRITE_ROLES:
        raise HTTPException(status_code=403, detail="Not authorized")
    await get_inquiry_or_404(db, inquiry_id)
    result = await db.execute(
        select(MenuSlot).where(MenuSlot.id == slot_id, MenuSlot.inquiry_id == inquiry_id)
    )
    slot = result.scalar_one_or_none()
    if not slot:
        raise HTTPException(status_code=404, detail="Menu slot not found")

    ext = os.path.splitext(file.filename or "")[1].lower()
    if ext not in settings.ALLOWED_EXTENSIONS:
        raise HTTPException(status_code=400, detail=f"File type '{ext}' not allowed")
    content = await file.read()
    if len(content) > settings.MAX_UPLOAD_SIZE:
        raise HTTPException(status_code=400, detail=f"File too large (max {settings.MAX_UPLOAD_SIZE // (1024 * 1024)}MB)")

    upload_dir = os.path.join(settings.UPLOAD_DIR, str(inquiry_id), "menu_slots")
    os.makedirs(upload_dir, exist_ok=True)
    file_path = os.path.join(upload_dir, f"slot{slot.slot_number}_{file.filename or 'unnamed'}")
    with open(file_path, "wb") as f:
        f.write(content)

    if slot.file_path and slot.file_path != file_path and os.path.isfile(slot.file_path):
        try:
            os.remove(slot.file_path)
        except OSError:
            pass
    slot.file_name = file.filename
    slot.file_path = file_path
    await db.commit()
    await db.refresh(slot)
    return MenuSlotResponse.model_validate(slot)


@router.patch("/{inquiry_id}/menu-slots/{slot_id}/final", response_model=MenuSlotResponse)
async def set_final_menu_slot(
    inquiry_id: uuid.UUID,
    slot_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role.name not in MENU_SLOT_WRITE_ROLES:
        raise HTTPException(status_code=403, detail="Not authorized")
    inquiry = await get_inquiry_or_404(db, inquiry_id)
    result = await db.execute(
        select(MenuSlot).where(MenuSlot.id == slot_id, MenuSlot.inquiry_id == inquiry_id)
    )
    slot = result.scalar_one_or_none()
    if not slot:
        raise HTTPException(status_code=404, detail="Menu slot not found")
    if not slot.file_name or not slot.file_path:
        raise HTTPException(status_code=400, detail="Upload a file to this slot before marking it final")
    await db.execute(
        update(MenuSlot).where(MenuSlot.inquiry_id == inquiry_id).values(is_final=False)
    )
    slot.is_final = True
    inquiry.menu_file_name = slot.file_name
    inquiry.menu_file_path = slot.file_path
    inquiry.menu_uploaded = True
    await db.flush()
    notify_result = await db.execute(
        select(User)
        .join(User.role)
        .where(Role.name == "sales_head", User.is_active.is_(True))
    )
    message = f"Final menu confirmed for {inquiry.client_name} ({inquiry.event_type})"
    for u in notify_result.scalars().all():
        db.add(Notification(
            user_id=u.id,
            title="Final menu confirmed",
            message=message[:500],
            type="menu",
            entity_type="inquiry",
            entity_id=inquiry.id,
        ))
    await db.commit()
    await db.refresh(slot)
    return MenuSlotResponse.model_validate(slot)


@router.delete("/{inquiry_id}/menu-slots/{slot_id}")
async def delete_menu_slot(
    inquiry_id: uuid.UUID,
    slot_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role.name not in MENU_SLOT_WRITE_ROLES:
        raise HTTPException(status_code=403, detail="Not authorized")
    result = await db.execute(
        select(MenuSlot).where(MenuSlot.id == slot_id, MenuSlot.inquiry_id == inquiry_id)
    )
    slot = result.scalar_one_or_none()
    if not slot:
        raise HTTPException(status_code=404, detail="Menu slot not found")
    if slot.file_path and os.path.isfile(slot.file_path):
        try:
            os.remove(slot.file_path)
        except OSError:
            pass
    await db.delete(slot)
    await db.commit()
    return {"message": "Menu slot deleted"}


@router.get("/{inquiry_id}/menu-slots/{slot_id}/download")
async def download_menu_slot_file(
    inquiry_id: uuid.UUID,
    slot_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    await get_inquiry_or_404(db, inquiry_id)
    result = await db.execute(
        select(MenuSlot).where(MenuSlot.id == slot_id, MenuSlot.inquiry_id == inquiry_id)
    )
    slot = result.scalar_one_or_none()
    if not slot or not slot.file_path or not os.path.isfile(slot.file_path):
        raise HTTPException(status_code=404, detail="No file uploaded for this slot")
    from fastapi.responses import FileResponse
    return FileResponse(
        path=slot.file_path,
        filename=slot.file_name,
        headers={"Cache-Control": "no-store, no-cache, must-revalidate"},
    )
