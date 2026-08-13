import csv
from openpyxl import load_workbook

ITEM_HEADER_WORDS = {"item", "item name", "item_name", "itemname", "product", "material", "ingredient", "name", "description"}


def read_file_rows(file_path: str, ext: str) -> list[list[str]]:
    rows: list[list[str]] = []
    if ext == ".csv":
        with open(file_path, "r", encoding="utf-8", errors="replace") as f:
            reader = csv.reader(f)
            for raw in reader:
                row = ["" if c is None else str(c).strip() for c in raw]
                if any(row):
                    rows.append(row)
    else:
        wb = load_workbook(file_path, data_only=True, read_only=True)
        try:
            ws = wb.active
            for raw in ws.iter_rows(values_only=True):
                row = ["" if c is None else (c if isinstance(c, (int, float)) else str(c).strip()) for c in raw]
                if any(row):
                    rows.append(row)
        finally:
            wb.close()
    return rows


def read_file_preview(file_path: str, ext: str) -> list[list]:
    """Read the complete file — no row or column caps."""
    if ext == ".csv":
        with open(file_path, "r", encoding="utf-8", errors="replace") as f:
            return [["" if c is None else str(c).strip() for c in raw] for raw in csv.reader(f) if any(raw)]
    wb = load_workbook(file_path, data_only=True, read_only=True)
    try:
        ws = wb.active
        return [
            ["" if c is None else (c if isinstance(c, (int, float)) else str(c).strip()) for c in raw]
            for raw in ws.iter_rows(values_only=True)
            if any(raw)
        ]
    finally:
        wb.close()


def parse_item_qty_file(file_path: str, ext: str) -> list[tuple[str, float, str | None]]:
    rows: list[tuple[str, float, str | None]] = []
    for row in read_file_rows(file_path, ext):
        item = row[0] if len(row) > 0 else ""
        if not item or item.lower() in ITEM_HEADER_WORDS:
            continue
        try:
            qty = float((row[1] if len(row) > 1 else "").replace(",", "")) if len(row) > 1 and row[1] else 0.0
        except (ValueError, TypeError):
            qty = 0.0
        if qty <= 0:
            continue
        unit = row[2] if len(row) > 2 and row[2] else None
        rows.append((item, qty, unit))
    return rows


def _to_float(value) -> float | None:
    if value is None or str(value).strip() == "":
        return None
    try:
        return float(str(value).replace(",", "").replace("₹", "").strip())
    except (ValueError, TypeError):
        return None


def _find_col(headers: list[str], keywords: set[str]) -> int | None:
    for i, h in enumerate(headers):
        if h and h.lower().strip() in keywords:
            return i
    return None


def parse_vendor_file(file_path: str, ext: str) -> list[dict]:
    rows = read_file_rows(file_path, ext)
    if not rows:
        return []
    headers = rows[0]
    name_col = _find_col(headers, {"vendor", "vendor name", "vendor_name", "supplier", "supplier name"})
    service_col = _find_col(headers, {"service", "service name", "service_name", "service type"})
    rate_col = _find_col(headers, {"rate", "price", "rate (rs)", "rate (inr)"})
    cost_col = _find_col(headers, {"total cost", "total", "total_cost", "cost", "amount"})
    remark_col = _find_col(headers, {"remark", "remarks", "note", "notes", "comments"})

    result = []
    for row in rows[1:]:
        def cell(i):
            return row[i] if i is not None and i < len(row) else ""
        vendor_name = str(cell(name_col)).strip() if name_col is not None else ""
        if not vendor_name:
            continue
        result.append({
            "vendor_name": vendor_name,
            "service_name": str(cell(service_col)).strip() or None if service_col is not None else None,
            "rate": _to_float(cell(rate_col)) if rate_col is not None else None,
            "total_cost": _to_float(cell(cost_col)) if cost_col is not None else None,
            "remark": str(cell(remark_col)).strip() or None if remark_col is not None else None,
        })
    return result


def parse_kitchen_inventory_file(file_path: str, ext: str) -> list[dict]:
    rows = read_file_rows(file_path, ext)
    if not rows:
        return []
    headers = rows[0]
    item_col = _find_col(headers, {"item", "item name", "item_name", "product", "dish", "preparation", "name"})
    prepared_col = _find_col(headers, {"prepared qty", "prepared", "prepared_qty", "qty prepared", "quantity prepared"})
    unit_col = _find_col(headers, {"unit", "uom"})
    used_col = _find_col(headers, {"used qty", "used", "used_qty", "qty used", "consumed"})
    remaining_col = _find_col(headers, {"remaining qty", "remaining", "remaining_qty", "qty remaining", "left"})
    remark_col = _find_col(headers, {"remark", "remarks", "note", "notes", "comments"})

    result = []
    for row in rows[1:]:
        def cell(i):
            return row[i] if i is not None and i < len(row) else ""
        item_name = str(cell(item_col)).strip() if item_col is not None else ""
        if not item_name:
            continue
        result.append({
            "item_name": item_name,
            "prepared_qty": _to_float(cell(prepared_col)) or 0 if prepared_col is not None else 0,
            "unit": str(cell(unit_col)).strip() or None if unit_col is not None else None,
            "used_qty": _to_float(cell(used_col)) or 0 if used_col is not None else 0,
            "remaining_qty": _to_float(cell(remaining_col)) or 0 if remaining_col is not None else 0,
            "remark": str(cell(remark_col)).strip() or None if remark_col is not None else None,
        })
    return result
